"""
export.py — V2.2 Excel report generation with OpenpyXL.

V2.2 changes
-------------
* "Alınan (₺)" column per shareholder (replaces "Durum" Ödendi/Ödenmedi).
* 3-colour fill: Green (paid), Yellow (kapora), Red (unpaid).
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import AnimalRecord, PaymentStatus

logger = logging.getLogger(__name__)

_FILL_GREEN = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
_FILL_RED = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
_FILL_HEADER = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
_FONT_HEADER = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_FONT_BODY = Font(name="Calibri", size=11)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

MAX_SHAREHOLDERS = 7


def _style(cell, *, font=_FONT_BODY, fill=None):
    cell.font = font
    cell.alignment = _CENTER
    cell.border = _BORDER
    if fill:
        cell.fill = fill


def generate_excel_report(records: List[AnimalRecord], dest: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Kurban Raporu"

    headers: List[str] = [
        "Hayvan ID", "Kesim Tarihi",
        "Toplam Fiyat (₺)", "Toplam Ağırlık (kg)", "Toplam Pay",
    ]
    for i in range(1, MAX_SHAREHOLDERS + 1):
        headers.append(f"Hissedar {i} Telefon")
        headers.append(f"Hissedar {i} Ad Soyad")
        headers.append(f"Hissedar {i} Pay")
        headers.append(f"Hissedar {i} Hisse Fiyat")
        headers.append(f"Hissedar {i} Alınan (₺)")
        headers.append(f"Hissedar {i} Durum")

    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        _style(cell, font=_FONT_HEADER, fill=_FILL_HEADER)

    for row_idx, animal in enumerate(records, start=2):
        c = ws.cell(row=row_idx, column=1, value=animal.animal_id)
        _style(c)
        c = ws.cell(row=row_idx, column=2, value=animal.slaughter_date.isoformat())
        _style(c)
        c = ws.cell(row=row_idx, column=3, value=str(animal.total_price))
        _style(c)
        c = ws.cell(row=row_idx, column=4, value=str(animal.total_weight))
        _style(c)
        c = ws.cell(row=row_idx, column=5, value=animal.total_fractions)
        _style(c)

        tf = animal.total_fractions
        for sh_idx, share in enumerate(animal.shares):
            base_col = 6 + sh_idx * 6  # 6 cols per shareholder now
            status = share.payment_status(animal.total_price, tf)
            if status == PaymentStatus.PAID:
                fill = _FILL_GREEN
                status_text = "Ödendi"
            elif status == PaymentStatus.PARTIAL:
                fill = _FILL_YELLOW
                status_text = "Kapora"
            else:
                fill = _FILL_RED
                status_text = "Ödenmedi"

            c = ws.cell(row=row_idx, column=base_col, value=share.phone)
            _style(c, fill=fill)
            c = ws.cell(row=row_idx, column=base_col + 1, value=share.shareholder_name)
            _style(c, fill=fill)
            c = ws.cell(row=row_idx, column=base_col + 2, value=share.share_fraction)
            _style(c, fill=fill)
            sh_price = share.price_for(animal.total_price, tf)
            c = ws.cell(row=row_idx, column=base_col + 3, value=str(sh_price))
            _style(c, fill=fill)
            c = ws.cell(row=row_idx, column=base_col + 4, value=str(share.paid_amount))
            _style(c, fill=fill)
            c = ws.cell(row=row_idx, column=base_col + 5, value=status_text)
            _style(c, fill=fill)

    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(records) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    ws.freeze_panes = "A2"
    wb.save(str(dest))
    logger.info("Excel report saved to %s (%d animals)", dest, len(records))
